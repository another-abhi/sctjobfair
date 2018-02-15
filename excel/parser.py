import openpyxl as xl
import openpyxl.utils as utl
import datetime
import MySQLdb

def test():
	workbook = xl.load_workbook('list.xlsx')
	print workbook.sheetnames
	sheetRanges = workbook['Sheet1']
	print sheetRanges.max_row
	print sheetRanges.max_column
	print utl.get_column_letter(1)
	print utl.column_index_from_string('AL')
	print sheetRanges['G2'].value

def parseExcel():
	fileName = raw_input("File name : ")
	workbook = xl.load_workbook(fileName)
	sheet = workbook['Sheet1']
	rowCount = sheet.max_row
	
	servername = "localhost";
	username = "ncscnlst_jobfair";
	password = "Adersh!23";
	dbname = "ncscnlst_jobfair18";
	
	db = MySQLdb.connect(servername, username, password, dbname)
	cursor = db.cursor()
	
	pid = ""
	fullName = ""
	contact = ""
	paymentStatus = ""
	paymentMethod = ""
	regMethod = ""
	companyCount = ""
	age = ""
	address = ""
	email = ""
	dob = ""
	gender = ""
	percentage10 = ""
	percentage12 = ""
	ugCourse = ""
	ugCollege = ""
	ugBranch = ""
	ugCgpa = ""
	backlogs = ""
	ugYop = ""
	fresher = ""
	experience = ""
	expCompany = ""
	pgCourse = ""
	pgCollege = ""
	pgCgpa = ""
	pgYop = ""
	company1 = ""
	company2 = "" 
	company3 = ""
	company4 = ""
	company5 = ""
	
	for currRow in range(2,rowCount+1):
		print '\n-----------------------------------------------------------------------------------\n'
		pid = sheet['B'+str(currRow)].value
		pid = int(pid)-100000
		print 'PID:' + str( pid ),
		
		fullName = sheet['C'+str(currRow)].value
		print ' Name:'+ fullName,
		
		email = sheet['D'+str(currRow)].value
		print ' Email:' + email,
		
		contact = sheet['E'+str(currRow)].value
		print ' Cont:' + contact,
		
		paymentStatus = sheet['G'+str(currRow)].value
		if( paymentStatus == 'SUCCESS' ):
			paymentStatus = 'paid'
		else:
			paymentStatus = 'notpaid'
		print ' PayStat:' + paymentStatus,
		
		companyCount = str(sheet['H'+str(currRow)].value)[0]
		print ' CmpCount:' + companyCount,
		
		age = sheet['J'+str(currRow)].value
		age = str(age).replace(" ", "")
		print ' Age:' + age,
		
		dob = sheet['K'+str(currRow)].value
		dob = datetime.datetime.strptime(dob, '%Y-%m-%dT%H:%M:%S.%fZ')
		print ' Dob:' + str(dob),

		address = sheet['L'+str(currRow)].value
		print ' Addr:' + address,
		
		gender = sheet['M'+str(currRow)].value
		print ' Gender:' + gender,
		
		percentage10 = sheet['N'+str(currRow)].value
		percentage10 = str(percentage10).replace(" ", "")
		if( percentage10 == 'O' ):
			percentage10='0'
		print ' P10:' + percentage10,
		
		percentage12 = sheet['O'+str(currRow)].value
		percentage12 = str(percentage12).replace(" ", "")
		if( percentage12 == 'O' ):
			percentage12='0'
		print ' P12:' + percentage12,
		
		ugCourse = sheet['P'+str(currRow)].value
		if( ugCourse == 'other' ):
			ugCourse = sheet['Q'+str(currRow)].value
		print ' UCourse:' + ugCourse,
		
		ugCollege = sheet['R'+str(currRow)].value
		if( ugCollege == 'other' ):
			ugCollege = sheet['S'+str(currRow)].value
		ugCollege = ugCollege.upper()
		print ' Ucollege:' + ugCollege,
		
		ugYop = sheet['T'+str(currRow)].value
		print ' Uyop:' + ugYop,
		
		ugCgpa = sheet['U'+str(currRow)].value
		ugCgpa = str(ugCgpa).replace(" ", "")
		if( ugCgpa == 'O' ):
			ugCgpa = '0'
		print ' Ugpa:' + ugCgpa,
		
		ugBranch = sheet['AK'+str(currRow)].value
		if( ugBranch == 'others' ):
			ugBranch = sheet['AL'+str(currRow)].value
		elif( ugBranch == '' or ugBranch == None ):
			ugBranch = 'nil'
		ugBranch = ugBranch.upper()
		print ' Ubranch:' + ugBranch,
		
		pgCourse = sheet['V'+str(currRow)].value
		if( pgCourse == 'other' ):
			pgCourse = sheet['W'+str(currRow)].value
		print ' Pcourse:' + pgCourse,
		
		if( pgCourse == 'none' ):
			pgCollege = 'nil'
			pgCgpa = '0'
			pgYop = 'nil'
		else:
			pgCollege = sheet['X'+str(currRow)].value
			if( pgCollege == 'other' ):
				pgCollege = sheet['Y'+str(currRow)].value
			pgCgpa = sheet['Z'+str(currRow)].value
			pgCgpa = str(pgCgpa).replace(" ", "")
			if( pgCgpa == 'nil' or pgCgpa == 'O' ):
				pgCgpa = '0'
			pgYop = sheet['AA'+str(currRow)].value
		
		print ' Pcollege:' + pgCollege,
		print ' Pgpa:' + pgCgpa,
		print ' Pyop:' + pgYop,
		
		backlogs = sheet['AB'+str(currRow)].value
		backlogs = str(backlogs).replace(" ", "")
		if( backlogs == '' or backlogs == None or backlogs == 'none' or backlogs== 'O'):
			backlogs = '0'
		print ' Backlog:' + backlogs,
		
		fresher = sheet['AH'+str(currRow)].value
		if( fresher == 'yes' ):
			fresher = 'y'
			experience = '0'
			expCompany = 'nil'
		elif( fresher == 'no' ):
			fresher = 'n'
			experience = sheet['AI'+str(currRow)].value
			experience = str(experience).replace(" ", "")
			expCompany = sheet['AJ'+str(currRow)].value
		
		if( paymentStatus == 'notpaid' ):
			paymentMethod = 'spot'
		elif( paymentStatus == 'paid'):
			paymentMethod = 'online'
		
		regMethod = 'online'
		
		company1 = sheet['AC'+str(currRow)].value
		company2 = sheet['AD'+str(currRow)].value
		company3 = sheet['AE'+str(currRow)].value
		company4 = sheet['AF'+str(currRow)].value
		company5 = sheet['AG'+str(currRow)].value
		print ' c1:' + company1,
		print ' c2:' + company2,
		print ' c3:' + company3,
		print ' c4:' + company4,
		print ' c5:' + company5
		
		sqlQuery = 'insert into participant(pid,paymentstatus,paymentmethod,regmethod,companycount,fullname,address,email,age,dob,contact,gender,percentage10,percentage12,ugcourse,ugcollege,ugbranch,ugcgpa,backlogs,ugyop,fresher,experience,expcompany,pgcourse,pgcollege,pgcgpa,pgyop) values ('
		sqlQuery += str(pid)+", \""+paymentStatus+"\", \""+paymentMethod+"\", \""+regMethod+"\", "+companyCount+","
		sqlQuery += " \""+fullName+"\", \""+address+"\", \""+email+"\", "+age+", \""+str(dob)+"\", \""+contact+"\", \""+gender+"\", "+percentage10+", "+percentage12+","
		sqlQuery += " \""+ugCourse+"\", \""+ugCollege+"\", \""+ugBranch+"\", "+ugCgpa+", "+backlogs+", \""+ugYop+"\", \""+fresher+"\", "+experience+", \""+expCompany+"\","
		sqlQuery += " \""+pgCourse+"\", \""+pgCollege+"\", "+pgCgpa+", \""+pgYop+"\" );"
	
		print sqlQuery
		cursor.execute(sqlQuery)
		if(company1 != 'nil'):
			sqlQuery = 'insert into participation values('+str(pid)+', \"'+company1+'\" );'
			print sqlQuery+"\n"
			cursor.execute(sqlQuery)
		if(company2 != 'nil'):
			sqlQuery = 'insert into participation values('+str(pid)+', \"'+company2+'\" );'
			print sqlQuery+"\n"
			cursor.execute(sqlQuery)
		if(company3 != 'nil'):
			sqlQuery = 'insert into participation values('+str(pid)+', \"'+company3+'\" );'
			print sqlQuery+"\n"
			cursor.execute(sqlQuery)
		if(company4 != 'nil'):
			sqlQuery = 'insert into participation values('+str(pid)+', \"'+company4+'\" );'
			print sqlQuery+"\n"
			cursor.execute(sqlQuery)
		if(company5 != 'nil'):
			sqlQuery = 'insert into participation values('+str(pid)+', \"'+company5+'\" );'
			print sqlQuery+"\n"
			cursor.execute(sqlQuery)
	option = raw_input("\nCommit changes (y/N): ")
	if option == 'y':
		db.commit()
		print 'databse committed..\n'
#test()
parseExcel()
